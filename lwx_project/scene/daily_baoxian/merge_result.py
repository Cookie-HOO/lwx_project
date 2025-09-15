"""

merge：
    用于将生成的df，合并到之前的信息中
    1. 在新的df中对字段进行抽取、整合到原有的excel
        - 原有的excel中有的字段都需要从新的df中抽取
        - 新的df中所有数据都需要放到原有的excel中（不需要进行任何的去重处理）
    2. 过期(获取招标文件的截止日期 超过今天)的条目删除，重新整理id
    3. 新加入的内容背景色标黄
        - 其他行本身有黄色背景需要去掉，只有新的行是黄色背景
    最后更新这个文件（写回）
    使用 openpyxl 进行处理

    原有的excel：注意标题在第三行
        序号,地区,采购方式,项目名称,采购单位名称,预算/限价（万元）,获取招标文件的截止日期,招采平台
    新的df：
        获取状态、是否选择、详情链接、项目名称、采购单位名称、预算/限价（万元）、获取招标文件的截止日期、原标题、地区、发布日期、招采平台、采购方式、详情信息、链接

send：
    发送邮件
"""
import re

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import datetime
import os

from lwx_project.scene.daily_baoxian.const import OLD_RESULT_PATH
from lwx_project.utils.date import cal_days_diff
from lwx_project.utils.mail import send_mail


def merge(new_df):
    old_excel_path = OLD_RESULT_PATH
    # 1. 过滤新数据
    new_df = new_df[new_df["是否选择"] == "True"].copy()
    new_df["招采平台"] = new_df.apply(
        lambda row: f"{row['招采平台']}：\n{row['链接']}"
        if pd.notna(row['招采平台']) and pd.notna(row['链接'])
        else row['招采平台'] if pd.notna(row['招采平台'])
        else row['链接'] if pd.notna(row['链接'])
        else None,
        axis=1  # 必须加 axis=1 才是按行处理
    )
    new_df["获取招标文件的截止日期"] = new_df["获取招标文件的截止日期"].apply(
        lambda x: cal_days_diff(end_date=datetime.datetime.strptime(x, "%Y/%m/%d")) if (pd.notna(x) and str(x)) else None
    )
    # 检查文件是否存在
    if not os.path.exists(old_excel_path):
        print(f"文件 {old_excel_path} 不存在")
        return
    
    try:
        # 2. 读取原有Excel文件
        wb = openpyxl.load_workbook(old_excel_path)
        ws = wb.active
        
        # 3. 从新的DataFrame中提取需要的字段
        # 定义需要的字段映射关系
        field_mapping = {
            "地区": "地区",
            "采购方式": "采购方式",
            "项目名称": "项目名称",
            "采购单位名称": "采购单位名称",
            "预算/限价（万元）": "预算/限价（万元）",
            "获取招标文件的截止日期": "获取招标文件的截止日期",
            "招采平台": "招采平台"
        }
        
        # 提取新数据
        new_data = []
        for _, row in new_df.iterrows():
            new_row = []
            for target_field, source_field in field_mapping.items():
                if source_field in new_df.columns:
                    new_row.append(row[source_field])
                else:
                    new_row.append("")
            new_data.append(new_row)
        
        # 4. 处理原数据，删除过期条目
        today = datetime.date.today()
        
        # 找到标题行位置（第三行，索引为2）
        header_row_index = 2  # 0-based index
        
        # 获取所有行数据
        # 读取单元格对象而非值，以获取原始日期类型
        all_rows = list(ws.iter_rows(min_row=header_row_index + 1))
        
        # 过滤过期条目
        valid_rows = []
        for row_cells in all_rows:
            # 将单元格对象转换为值列表
            row = [cell.value for cell in row_cells]
            # 获取日期单元格的值（可能是datetime对象或字符串）
            if len(row_cells) > 6 and row_cells[6].value is not None:
                date_cell = row_cells[6]
                date_value = date_cell.value  # 确保有截止日期字段且不为空
                # 尝试解析日期
                try:
                    deadline = None
                    # 处理Excel日期格式和不同字符串格式
                    # 处理datetime对象（Excel日期类型）
                    if isinstance(date_value, datetime.datetime):
                        deadline = date_value.date()
                    # 处理字符串格式
                    elif isinstance(date_value, str):
                        date_str = date_value.strip()
                        # 尝试多种日期格式
                        date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%m/%d/%Y', '%d/%m/%Y', '%Y.%m.%d', '%m月%d日']
                        for fmt in date_formats:
                            try:
                                deadline = datetime.datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        # 如果仍无法解析且格式为'%m月%d日'，补充当前年份
                        if deadline is None and re.match(r'^\d+月\d+日$', date_str):
                            current_year = datetime.date.today().year
                            try:
                                deadline = datetime.datetime.strptime(f'{current_year}年{date_str}', '%Y年%m月%d日').date()
                            except ValueError:
                                pass
                    # 处理Excel序列号日期（如果是数字）
                    elif isinstance(date_value, (int, float)):
                        try:
                            # Excel日期序列号转换（1900-01-01为1）
                            deadline = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=date_value)
                            deadline = deadline.date()
                        except Exception:
                            pass
                    
                    # 检查是否过期
                    if deadline is None or deadline >= today:
                        valid_rows.append(list(row[1:]))  # 去掉序号列
                except Exception:
                    # 如果日期解析失败，保留该行
                    valid_rows.append(list(row[1:]))
            else:
                # 如果没有截止日期字段或为空，保留该行
                valid_rows.append(list(row[1:]))
        
        # 5. 合并新旧数据
        combined_data = valid_rows + new_data

        # 6. 彻底删除标题行之后的所有数据行（保留标题）
        start_delete_row = header_row_index + 1  # 第4行开始是数据

        # 获取当前最大行号
        max_row = ws.max_row
        if max_row >= start_delete_row:
            # 删除从 start_delete_row 到最后一行的所有行
            ws.delete_rows(start_delete_row, max_row - start_delete_row + 1)

        # 7. 写入合并后的数据并添加序号
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        simsun_font = Font(name='SimSun', size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 从 header_row_index + 1 开始写入新数据（现在这里是第一行数据）
        for idx, row_data in enumerate(combined_data, 1):
            # 写入序号（第1列）
            serial_cell = ws.cell(row=header_row_index + idx, column=1)
            serial_cell.value = idx
            serial_cell.alignment = center_align
            serial_cell.border = border

            # 新增行标记：如果 idx > len(valid_rows)，说明是本次新增的
            if idx > len(valid_rows):
                serial_cell.fill = yellow_fill

            # 设置行高
            ws.row_dimensions[header_row_index + idx].height = 30

            # 写入其他字段（从第2列开始）
            for col_idx, value in enumerate(row_data, 2):
                cell = ws.cell(row=header_row_index + idx, column=col_idx)
                cell.value = value
                cell.font = simsun_font

                # 格式化“获取招标文件的截止日期”列（第7列）
                if col_idx == 7:
                    cell.number_format = 'm"月"d"日"'

                # 招采平台列（第8列）自动换行
                if col_idx == 8:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = center_align

                # 只有非空值才加边框
                if value is not None and str(value).strip() != '':
                    cell.border = border

                # 新增行设置黄色背景
                if idx > len(valid_rows):
                    cell.fill = yellow_fill
                    if value is not None and str(value).strip() != '':
                        cell.border = border  # 非空才加边框

        # 8. 清理后续可能残留的空行（保险起见）
        last_written_row = header_row_index + len(combined_data)
        max_existing_row = ws.max_row
        if max_existing_row > last_written_row:
            ws.delete_rows(last_written_row + 1, max_existing_row - last_written_row)
            
        # 9. 保存文件
        wb.save(old_excel_path)
        print(f"成功将新数据合并到 {old_excel_path}")
        
    except Exception as e:
        print(f"处理Excel文件时出错：{str(e)}")


def send():
    from_email = "liwenxuan_0112@126.com"
    to_email = "liwenxuanrs@abchina.com"
    subject = "每日招标收集"
    attachments = [OLD_RESULT_PATH]
    send_mail(
        from_email=from_email,
        to_email=to_email,
        subject=subject,
        body="",
        attachments=attachments
    )

if __name__ == '__main__':
    df_ = pd.DataFrame({
        "获取状态": ["✅"],
        "是否选择": ["True"],
        "详情链接": [""],
        "项目名称": ["好项目"],
        "采购单位名称": ["北京"],
        "预算/限价（万元）": ["10"],
        "获取招标文件的截止日期": ["2025/09/05"],
        "原标题": ["一个好项目"],
        "地区": ["北京"],
        "发布日期": ["asdf"],
        "招采平台": ["中国政府网"],
        "采购方式": ["公开招标"],
        "详情信息": ["sdfdsaf"],
        "链接": ["https://www.baidu.com"],
    })
    merge(df_)
