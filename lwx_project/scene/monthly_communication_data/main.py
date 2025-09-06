"""
每月做一次
每月需要导出上一个月的数据表A（10mb左右）
基于表A计算一个截止到这个月的全年数据


当前25年8月
场景1: 初始化
    传1-8月的+人员数据，进行初始化，计算8月的
场景2: 之前某个月传错了
    传3月的，此时可传可不传人员数据，重新计算截止3月、4月、5月 到 8月的
场景3: 正常
    传8月的+人员数据，基于之前的1-7月的结果，计算截止8月的

之前算过的，加上新上传的，必须是连续的，然后计算最后一个月的
"""
import datetime
import os
import typing

import pandas as pd

from lwx_project.scene.monthly_communication_data.const import DETAIL_PATH, IMPORTANT_PATH
from lwx_project.utils.time_cost import time_cost

UPLOAD_TYPE_OFFICER = "内外勤人数数据"
UPLOAD_TYPE_TUANXIAN = "团险核心数据"

class UploadInfo:
    def __init__(self, year,  upload_tuanxian_month_dict, important_month_dict, officer_path):
        self.year = year
        self.upload_tuanxian_month_dict = upload_tuanxian_month_dict
        self.important_month_dict = important_month_dict

        self.officer_path = officer_path
        pass



def check_upload(upload_file_list) -> (bool, str, typing.Optional[UploadInfo]):
    """校验上传的文件
    1. 1个或多个「团险核心数据」 + 0个或1个 「内外勤人数数据」
    2. 判断
        两种类型的文件数量校验
        多个团险核心数据的年份一致
        多个团险核心数据的月份，和important中做完的月份，要连贯，从1月开始
        如果要做新的（上传的团险核心数据中的月份比important中的月份大），必须要有「内外勤人数数据」
    """
    # 1. 区分detail 文件和统计文件
    detail_path_list = []  # 团限核心数据
    officer_path_list = []  # 内勤外勤人员数据
    for path in upload_file_list:
        file_type = get_upload_type(path)
        if file_type == UPLOAD_TYPE_OFFICER:
            officer_path_list.append(path)
        elif file_type == UPLOAD_TYPE_TUANXIAN:
            detail_path_list.append(path)
    # 2. 校验
    # 2.1 长度校验
    if len(officer_path_list) > 1:
        return False, "最多支持一个内外勤人员数据", None
    if len(detail_path_list) == 0:
        return False, "必须上传团险核心数据", None
    # 2.2 校验所有的年份一致
    date_list = [get_month_from_detail_path(detail_path) for detail_path in detail_path_list]
    year_list = [d.year for d in date_list]
    month_list = [d.month for d in date_list]
    if len(set(year_list)) != 1:
        return False, f"上传的团险核心数据中，存在不一致的年份：{set(year_list)}", None # 存在年份不一致的数据
    year = year_list[0]  # noqa: Unexpected Type

    # 2.3 校验和important内容是否连续
    # 先确保important目录下一定有这个年份的文件夹
    target_year_dir = os.path.join(IMPORTANT_PATH, str(year))
    os.makedirs(target_year_dir, exist_ok=True)
    # 找到所有important当前年中的所有已做完的月
    important_month_list = [int(file_name.split("月")[0]) for file_name in os.listdir(IMPORTANT_PATH)]  # "12月同业交流数据.xlsx"
    # 确保 month_list + important_month_list 一定是从1开始的连续数字（可重复）
    all_month_list = month_list + important_month_list
    for i in range(1, 13):  # noqa: Unexpected Type
        if i > len(set(all_month_list)):  # 一旦超过整体长度，说明校验成功
            break
        if i not in month_list + important_month_list:  # 如果没超过整体长度，且没有
            return False, f"整体月份不连续，缺少{i}月团险核心数据", None

    # 2.4 校验如果是新作的内容，必须包含内勤外勤的人员数据
    if max(month_list) > max(important_month_list) and len(officer_path_list) == 0:
        return False, "上传新月份时必须提供内外勤人员数据", None
    return True, "", UploadInfo(
        year=year,
        upload_tuanxian_month_dict={k: v for k, v in zip(month_list, detail_path_list)},
        important_month_dict={k: os.path.join(target_year_dir, v) for k, v in zip(important_month_list, os.listdir(IMPORTANT_PATH))},
        officer_path=officer_path_list[0] if officer_path_list else None
    )

@time_cost
def get_upload_type(upload_file_path: str) -> str:
    from openpyxl import load_workbook

    wb = None
    try:
        # 直接调用，不使用 with
        wb = load_workbook(upload_file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel 文件没有活动的工作表")
        header_names = [cell.value for cell in ws[1]]
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件 '{upload_file_path}': {e}")
    finally:
        if wb is not None:
            wb.close()  # 手动关闭文件，释放资源

    for col_name in header_names:
        if col_name is not None and "机构代码" in str(col_name):
            return UPLOAD_TYPE_TUANXIAN

    return UPLOAD_TYPE_OFFICER

@time_cost
def get_month_from_detail_path(detail_path: str) -> datetime.date:
    """
    高效读取 Excel 文件前几行，提取「日期」列的第一个值，返回 date 对象。
    """
    try:
        # 只读前1行数据（加上列名，共2行）
        df = pd.read_excel(detail_path, nrows=1, dtype=str)
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件: {e}")

    if df.empty:
        raise ValueError("Excel 文件为空")

    # 查找「日期」列
    date_col = None
    for col in df.columns:
        if str(col).strip() in ["日期", "date", "Date", "DATE"]:
            date_col = col
            break

    if date_col is None:
        raise ValueError(f"未找到「日期」列，列名为: {list(df.columns)}")

    # 获取第一行的值
    first_value = df.iloc[0][date_col]
    if pd.isna(first_value) or not str(first_value).strip():
        raise ValueError("「日期」列第一行为空")

    date_str = str(first_value).strip()

    # 解析日期
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y.%m.%d"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    # 最后尝试用 pandas 自动解析
    try:
        parsed = pd.to_datetime(date_str)
        return parsed.date()
    except:
        raise ValueError(f"无法解析日期字符串: {date_str}")



if __name__ == '__main__':
    is_success, error_msg, res = check_upload([
        DETAIL_PATH
    ])
    print()