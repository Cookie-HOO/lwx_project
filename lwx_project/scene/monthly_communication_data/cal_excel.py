"""
在check_excel之后，输入是 UploadInfo 对象

最核心的部分是，给定一个detail的path，计算所有的分组汇总结果
"""
import typing

import pandas as pd
import os

from openpyxl import load_workbook

from lwx_project.scene.monthly_communication_data.check_excel import UploadInfo
from lwx_project.scene.monthly_communication_data.const import DETAIL_PATH, IMPORTANT_PATH, TEMPLATE_PATH
from lwx_project.utils.file import copy_file
from lwx_project.utils.time_cost import time_cost


def cal_and_merge(upload_info: UploadInfo, code_rules_dict, after_one_done_callback):
    # 1. 计算所有的结果
    caled_result_list = detail_group_by(
        upload_info=upload_info,
        code_map=code_rules_dict,
        after_one_done_callback=after_one_done_callback,
    )

    # 2. 填充指定位置
    files_map = merge_caled_result(upload_info, caled_result_list)
    return files_map


@time_cost
def fill_current_month_data(target_year_dir, upload_info: UploadInfo, result_list):

    caled_paths = []

    # 将计算结果填充模板
    for month, result_df in zip(upload_info.upload_tuanxian_month_dict.keys(), result_list):
        this_path = os.path.join(target_year_dir, str(month) + "月同业交流数据.xlsx")
        # 拷贝模板
        copy_file(TEMPLATE_PATH, this_path)

        # 使用openpyxl来操作Excel，保留原有格式
        try:
            # 加载工作簿
            wb = load_workbook(this_path)
            # 获取第一个工作表
            ws = wb.active

            # 创建分公司到行号的映射
            company_to_row = {}
            # 从第4行开始查找分公司（模板前3行是标题）
            for row in range(4, ws.max_row + 1):
                company_cell = ws.cell(row=row, column=1)
                company_name = company_cell.value
                if company_name and isinstance(company_name, str):
                    company_to_row[company_name] = row

            # 将result_df中的数据填充到模板中
            for _, row_data in result_df.iterrows():
                company_name = row_data['分公司']
                if company_name in company_to_row:
                    row_num = company_to_row[company_name]

                    # 填充数据，保留原有格式
                    # 短期：<= 1年
                    # 意外险 - 第2列
                    ws.cell(row=row_num, column=2).value = row_data['意外险']
                    # 健康险 - 第3列
                    ws.cell(row=row_num, column=3).value = row_data['健康险']
                    # 寿险 - 第4列
                    ws.cell(row=row_num, column=4).value = row_data['寿险']
                    # 医疗基金 - 第5列
                    ws.cell(row=row_num, column=5).value = row_data['医疗基金']
                    # 合计 - 第6列
                    ws.cell(row=row_num, column=6).value = row_data['合计']
                    # 长期：>1年
                    # 长期意外	长期健康险	长期寿险 都是0（12、13、14是当月，17、18、19是累计全年
                    ws.cell(row=row_num, column=12).value = 0
                    ws.cell(row=row_num, column=13).value = 0
                    ws.cell(row=row_num, column=14).value = 0
                    ws.cell(row=row_num, column=17).value = 0
                    ws.cell(row=row_num, column=18).value = 0
                    ws.cell(row=row_num, column=19).value = 0

                    # 年金险 - 第15列和第16列
                    ws.cell(row=row_num, column=15).value = row_data['年金险']
                    ws.cell(row=row_num, column=16).value = row_data['年金险']

            # 保存工作簿
            wb.save(this_path)
            caled_paths.append(this_path)
            print(f"成功填充 {this_path} 的数据")
        except Exception as e:
            print(f"填充 {this_path} 的数据时出错: {str(e)}")
            continue

@time_cost
def fill_cumulative_data(target_year_dir, upload_info: UploadInfo):
    # 将汇总结果填充模板
    need_cal_min_month = min(upload_info.upload_tuanxian_month_dict.keys())  # 最小需要计算的月
    candidate_files = [i for i in os.listdir(target_year_dir) if i.endswith(".xlsx")]
    files = sorted(candidate_files, key=lambda x: int(x.split("月")[0]))
    files_map = {f: os.path.join(target_year_dir, f) for f in files}
    # 存储每个月的数据，用于计算后续月份的累计值
    month_data_dict = {}

    for file, this_path in files_map.items():  # 按顺序处理文件
        this_parts = file.split("月")
        this_month = int(this_parts[0])
        path_suffix = "月" + this_parts[1]

        # 遍历所有important中的文件（该算当月的已经保存到对应文件中了）
        # 只需要从上传的最小月开始计算汇总，但是最小的月，需要上一个月的汇总
        if this_month < need_cal_min_month-1:  # 需要额外多计算一个月份，用于给第一个月份做铺垫（第一个月份也需要上一个月份的汇总）
            continue
        # 加载当前文件
        try:
            wb = load_workbook(this_path)
            ws = wb.active
            MAX_ROW = ws.max_row - 1  # 最后一行是总计，不需要处理

            if this_month == 1:  # 1月的情况，汇总值就是当月值
                # 从第4行到第26行（包含）
                for row_num in range(4, MAX_ROW + 1):
                    # 短期部分：第2、3、4、5、6列分别对应7、8、9、10、11列
                    for col_idx in range(2, 7):
                        current_value = ws.cell(row=row_num, column=col_idx).value or 0
                        # 确保值是数值类型
                        current_value = float(current_value) if isinstance(current_value, (int, float, str)) else 0
                        if isinstance(current_value, str):
                            try:
                                current_value = float(current_value)
                            except ValueError:
                                current_value = 0
                        ws.cell(row=row_num, column=col_idx + 5).value = current_value  # 短期全年（+5列）

                    # 长期部分：第12、13、14、15、16列分别对应17、18、19、20、21列
                    for col_idx in range(12, 17):
                        current_value = ws.cell(row=row_num, column=col_idx).value or 0
                        # 确保值是数值类型
                        current_value = float(current_value) if isinstance(current_value, (int, float, str)) else 0
                        if isinstance(current_value, str):
                            try:
                                current_value = float(current_value)
                            except ValueError:
                                current_value = 0
                        ws.cell(row=row_num, column=col_idx + 5).value = current_value  # 长期全年（+5列）

                # 保存工作簿
                wb.save(this_path)

                # 保存1月的数据供后续月份使用
                month_data = {
                    'short_term': {},  # 存储短期全年数据
                    'long_term': {}  # 存储长期全年数据
                }
                for row_num in range(4, MAX_ROW + 1):
                    month_data['short_term'][row_num] = [ws.cell(row=row_num, column=col).value or 0 for col in
                                                         range(7, 12)]
                    month_data['long_term'][row_num] = [ws.cell(row=row_num, column=col).value or 0 for col in
                                                        range(17, 22)]
                month_data_dict[this_month] = month_data

            else:  # 非1月的情况，汇总值是上月汇总值加本月当月值
                # 找到上个月的文件
                last_month = this_month - 1
                if last_month in month_data_dict:
                    last_month_data = month_data_dict[last_month]

                    # 从第4行到第26行（包含）
                    for row_num in range(4, MAX_ROW + 1):
                        # 短期部分：上月汇总值（7、8、9、10、11）+ 本月当月值（2、3、4、5、6）
                        for col_idx in range(5):
                            last_value = last_month_data['short_term'].get(row_num, [0] * 5)[col_idx] or 0
                            current_value = ws.cell(row=row_num, column=col_idx + 2).value or 0

                            # 确保值是数值类型
                            last_value = float(last_value) if isinstance(last_value, (int, float, str)) else 0
                            current_value = float(current_value) if isinstance(current_value,
                                                                               (int, float, str)) else 0
                            if isinstance(last_value, str):
                                try:
                                    last_value = float(last_value)
                                except ValueError:
                                    last_value = 0
                            if isinstance(current_value, str):
                                try:
                                    current_value = float(current_value)
                                except ValueError:
                                    current_value = 0

                            # 计算并更新汇总值
                            total_value = last_value + current_value
                            ws.cell(row=row_num, column=col_idx + 7).value = total_value  # 短期全年（7-11列）

                        # 长期部分：上月汇总值（17、18、19、20、21）+ 本月当月值（12、13、14、15、16）
                        for col_idx in range(5):
                            last_value = last_month_data['long_term'].get(row_num, [0] * 5)[col_idx] or 0
                            current_value = ws.cell(row=row_num, column=col_idx + 12).value or 0

                            # 确保值是数值类型
                            last_value = float(last_value) if isinstance(last_value, (int, float, str)) else 0
                            current_value = float(current_value) if isinstance(current_value,
                                                                               (int, float, str)) else 0
                            if isinstance(last_value, str):
                                try:
                                    last_value = float(last_value)
                                except ValueError:
                                    last_value = 0
                            if isinstance(current_value, str):
                                try:
                                    current_value = float(current_value)
                                except ValueError:
                                    current_value = 0

                            # 计算并更新汇总值
                            total_value = last_value + current_value
                            ws.cell(row=row_num, column=col_idx + 17).value = total_value  # 长期全年（17-21列）

                    # 保存工作簿
                    wb.save(this_path)

                # 保存当前月份的数据供后续月份使用
                month_data = {
                    'short_term': {},
                    'long_term': {}
                }
                for row_num in range(4, MAX_ROW + 1):
                    month_data['short_term'][row_num] = [ws.cell(row=row_num, column=col).value or 0 for col in
                                                         range(7, 12)]
                    month_data['long_term'][row_num] = [ws.cell(row=row_num, column=col).value or 0 for col in
                                                        range(17, 22)]
                month_data_dict[this_month] = month_data
        except Exception as e:
            print(f"处理 {this_path} 的汇总值时出错: {str(e)}")
            continue

    return files_map

# 核心的分组计算过程
@time_cost
def detail_group_by(upload_info: UploadInfo, code_map: dict, after_one_done_callback: typing.Callable[[int], None]=None):
    """
    upload_info:
    code_map: 计算汇总结果时需要忽略或增加的代码，负数为忽略，正数为增加，绝对值为代码
        {
            "意外险": [],
            "健康险": [-7824],
            "寿险": [],
            "医疗基金": [+7824, +7854],
            "年金险": [-2801]
        }

    针对每一个excel执行下列过程（注意这里的每个excel可能在10mb左右，需要注意性能）：
        1. 读取第一个sheet
        2. 仅保留三列（第一行就是列名）
            ”机构名称”、“险种代码”、“保费收入/支出（含税）”
        3. 提取结构名称的前两个字符，如果前两个字符为“黑龙”，就补成 “黑龙江”，变成“分公司”列
            此时可以删掉机构名称列
        4. 按照“分公司”列进行groupby计算
            意外险：险种代码 以68开头 的行
            健康险：险种代码 以78开头，但是不包括7824 的行
                这里的7824是由 code_map 变量决定的，动态传入的，7824只是举个例子
            寿险: 意外险：险种代码 以18开头 的行
            医疗基金：险种代码 是7824 的行
                这里的7824是由 code_map 变量决定的，动态传入的，7824只是举个例子
            合计：上述四种保险的合计
            计算上述四种保险的价格：即 求和 保费收入/支出（含税） 列
                即每个分公司需要按照对应条件筛选后，对 保费收入/支出（含税） 列求和

            再增加一个年金险（算合计时不考虑）
                28开头，不考虑2801
        5. 最终生成的表格就是，第一列是groupby的分公司，后面6列是对应的四种保险的汇总以及合计 再加年金险
            共7列
    每个excel返回6列的一个dataframe
    """
    result_dfs = []
    excel_path_list = upload_info.upload_tuanxian_month_dict.values()
    # 验证excel_path_list不为空
    if not excel_path_list:
        return result_dfs
    
    # 验证code_map的基本结构
    required_keys = ["意外险", "健康险", "寿险", "医疗基金"]
    for key in required_keys:
        if key not in code_map:
            code_map[key] = []
    
    for index, excel_path in enumerate(excel_path_list):
        try:
            # 检查文件是否存在
            if not os.path.exists(excel_path):
                print(f"警告：文件不存在: {excel_path}")
                continue
            
            # 检查文件是否为Excel文件
            if not (excel_path.endswith('.xlsx') or excel_path.endswith('.xls')):
                print(f"警告：文件不是Excel文件: {excel_path}")
                continue
            
            # 读取Excel文件的第一个sheet，优化性能
            df = pd.read_excel(excel_path, sheet_name=0, engine='openpyxl', usecols=None)
            
            # 只保留需要的三列
            columns_to_keep = [
                '机构名称',  # 可能需要处理全角引号
                '险种代码',
                '保费收入/支出（含税）'
            ]
            
            # 处理可能的列名不一致问题
            actual_columns = []
            for col in columns_to_keep:
                found = False
                for actual_col in df.columns:
                    # 处理可能的全角字符和空格问题
                    clean_actual_col = ''.join(char for char in str(actual_col) if char.isprintable()).strip()
                    clean_target_col = ''.join(char for char in col if char.isprintable()).strip()
                    
                    if clean_target_col in clean_actual_col or clean_actual_col in clean_target_col:
                        actual_columns.append(actual_col)
                        found = True
                        break
                if not found:
                    # 如果找不到完全匹配的列，尝试找到最相似的
                    for actual_col in df.columns:
                        clean_actual_col = ''.join(char for char in str(actual_col) if char.isprintable()).strip()
                        if any(keyword in clean_actual_col for keyword in ['机构', '险种', '保费']):
                            actual_columns.append(actual_col)
                            found = True
                            break
                if not found:
                    raise ValueError(f"在Excel文件中找不到列: {col}")
            
            # 重命名列以便后续处理
            df = df[actual_columns].copy()
            df.columns = columns_to_keep
            
            # 创建"分公司"列
            def process_company_name(name):
                if isinstance(name, str):
                    if len(name) >= 2:
                        first_two = name[:2]
                        if first_two == '黑龙':
                            return '黑龙江'
                        return first_two
                return str(name)[:2] if isinstance(name, str) else str(name)
            
            df['分公司'] = df['机构名称'].apply(process_company_name)
            
            # 移除"机构名称"列
            df = df.drop('机构名称', axis=1)
            
            # 确保"险种代码"是字符串类型
            df['险种代码'] = df['险种代码'].astype(str)
            
            # 确保"保费收入/支出（含税）"是数值类型
            df['保费收入/支出（含税）'] = pd.to_numeric(df['保费收入/支出（含税）'], errors='coerce')
            df = df.dropna(subset=['保费收入/支出（含税）'])
            
            # 准备分组计算
            grouped = df.groupby('分公司')
            result_df = pd.DataFrame()
            result_df['分公司'] = list(grouped.groups.keys())
            
            # 计算各种保险类型的保费总和
            
            # 1. 意外险计算
            # 基础条件：险种代码必须以68开头
            # 从code_map['意外险']中获取所有正数代码（表示额外包含的代码）和负数代码（表示要排除的代码）
            include_accident_codes = [str(abs(code)) for code in code_map.get('意外险', []) if code > 0]
            exclude_accident_codes = [str(abs(code)) for code in code_map.get('意外险', []) if code < 0]
            
            accident_insurance = grouped.apply(
                lambda x: x[
                    # 必须满足基础条件：以68开头
                    (x['险种代码'].str.startswith('68', na=False)) & 
                    # 额外包含的代码（如果有）
                    ((len(include_accident_codes) == 0) | (x['险种代码'].isin(include_accident_codes))) & 
                    # 排除指定的代码
                    (~x['险种代码'].isin(exclude_accident_codes))
                ]['保费收入/支出（含税）'].sum()
            )
            
            # 2. 健康险计算
            # 基础条件：险种代码必须以78开头
            # 从code_map['健康险']中获取所有正数代码（表示额外包含的代码）和负数代码（表示要排除的代码）
            include_health_codes = [str(abs(code)) for code in code_map.get('健康险', []) if code > 0]
            exclude_health_codes = [str(abs(code)) for code in code_map.get('健康险', []) if code < 0]
            
            health_insurance = grouped.apply(
                lambda x: x[
                    # 必须满足基础条件：以78开头
                    (x['险种代码'].str.startswith('78', na=False)) & 
                    # 额外包含的代码（如果有）
                    ((len(include_health_codes) == 0) | (x['险种代码'].isin(include_health_codes))) & 
                    # 排除指定的代码
                    (~x['险种代码'].isin(exclude_health_codes))
                ]['保费收入/支出（含税）'].sum()
            )
            
            # 3. 寿险计算
            # 基础条件：险种代码必须以18开头
            # 从code_map['寿险']中获取所有正数代码（表示额外包含的代码）和负数代码（表示要排除的代码）
            include_life_codes = [str(abs(code)) for code in code_map.get('寿险', []) if code > 0]
            exclude_life_codes = [str(abs(code)) for code in code_map.get('寿险', []) if code < 0]
            
            life_insurance = grouped.apply(
                lambda x: x[
                    # 必须满足基础条件：以18开头
                    (x['险种代码'].str.startswith('18', na=False)) & 
                    # 额外包含的代码（如果有）
                    ((len(include_life_codes) == 0) | (x['险种代码'].isin(include_life_codes))) & 
                    # 排除指定的代码
                    (~x['险种代码'].isin(exclude_life_codes))
                ]['保费收入/支出（含税）'].sum()
            )
            
            # 4. 医疗基金计算
            # 从code_map['医疗基金']中获取所有正数代码（表示要包含的代码）和负数代码（表示要排除的代码）
            include_medical_codes = [str(abs(code)) for code in code_map.get('医疗基金', []) if code > 0]
            exclude_medical_codes = [str(abs(code)) for code in code_map.get('医疗基金', []) if code < 0]
            
            medical_fund = grouped.apply(
                lambda x: x[
                    # 必须满足包含条件
                    (x['险种代码'].isin(include_medical_codes)) & 
                    # 排除指定的代码
                    (~x['险种代码'].isin(exclude_medical_codes))
                ]['保费收入/支出（含税）'].sum()
            )
            
            # 计算合计
            total = accident_insurance + health_insurance + life_insurance + medical_fund

            # 5. 年金险（需要单独算合计，不和其他类型一起）
            include_annu_codes = [str(abs(code)) for code in code_map.get('年金险', []) if code > 0]
            exclude_annu_codes = [str(abs(code)) for code in code_map.get('年金险', []) if code < 0]

            annu_fund = grouped.apply(
                lambda x: x[
                    # 必须满足基础条件：以28开头
                    (x['险种代码'].str.startswith('28', na=False)) &
                    # 额外包含的代码（如果有）
                    ((len(include_annu_codes) == 0) | (x['险种代码'].isin(include_annu_codes))) &
                    # 排除指定的代码
                    (~x['险种代码'].isin(exclude_annu_codes))
                    ]['保费收入/支出（含税）'].sum()
            )
            
            # 将计算结果添加到结果DataFrame
            result_df = result_df.set_index('分公司')
            result_df['意外险'] = accident_insurance
            result_df['健康险'] = health_insurance
            result_df['寿险'] = life_insurance
            result_df['医疗基金'] = medical_fund
            result_df['合计'] = total
            result_df['年金险'] = annu_fund

            # 重置索引，使分公司成为普通列
            result_df = result_df.reset_index()
            
            # 填充可能的NaN值为0
            result_df = result_df.fillna(0)
            
            result_dfs.append(result_df)
        except Exception as e:
            print(f"处理文件 {excel_path} 时出错: {str(e)}")
            continue
        finally:
            if after_one_done_callback is not None:
                after_one_done_callback(index)
    return result_dfs

# 核心的汇总过程
@time_cost
def merge_caled_result(upload_info: UploadInfo, result_list):
    # 1 找到对应的模板，如果没有拷贝一个新的出来
    # 2 将结果逐一填充进去
    # 返回所有可供下载的文件名称和路径的映射

    target_year_dir = os.path.join(IMPORTANT_PATH, str(upload_info.year))

    # 确保目标目录存在
    os.makedirs(target_year_dir, exist_ok=True)
    fill_current_month_data(target_year_dir, upload_info, result_list)  # 填充当月数据

    files_map = fill_cumulative_data(target_year_dir, upload_info)  # 填充累计数据
    return files_map

if __name__ == '__main__':
    upload_info_ = UploadInfo(
        year=2025,
        upload_tuanxian_month_dict = {3: DETAIL_PATH, 4: DETAIL_PATH, 5: DETAIL_PATH},
        important_month_dict={},
        officer_path = None,
    )

    result_list_ = detail_group_by(upload_info_,{
            "意外险": [],
            "健康险": [-7824, -7854],  # 后面可能动态变
            "寿险": [],
            "医疗基金": [+7824, +7854],  # 后面可能动态变
            "年金险": [-2801],
        })

    merge_caled_result(upload_info_, result_list_)
    print()