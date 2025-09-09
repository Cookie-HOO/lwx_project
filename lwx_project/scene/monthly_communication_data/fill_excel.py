"""
在check_excel之后，输入是 UploadInfo 对象

最核心的部分是，给定一个detail的path，计算所有的分组汇总结果
"""
import os

from openpyxl import load_workbook

from lwx_project.scene.monthly_communication_data.check_excel import UploadInfo
from lwx_project.scene.monthly_communication_data.const import DETAIL_PATH, IMPORTANT_PATH, TEMPLATE_PATH, CALED_FILE
from lwx_project.utils.file import copy_file
from lwx_project.utils.time_cost import time_cost

# 模板相关信息
TEMPLATE_MIN_ROW = 6  # 从第6行开始
TEMPLATE_MAX_ROW = lambda ws: ws.max_row - 1  # 去掉合计
TEMPLATE_BAOXIAN_COL_NUM_MAP = {  # key：result_list 中的列，value：template中的列号
    # 意外险	健康险	寿险	医疗基金	合计
    "意外险": 4,
    "健康险": 5,
    "寿险": 6,
    "医疗基金": 7,
    "短期合计": 8,

    "年金险": 17,
    "长期合计": 18,
}
TEMPLATE_BAOXIAN_OMIT_COL_NUM_LIST = [14, 15, 16, 19, 20, 21]  # 分别是：长期意外	长期健康险	长期寿险的当月和全年

TEMPLATE_CURRENT_TO_CUMULATIVE_MAP = {  # 模板中当前列和汇总列的映射关系
    4: 9,  # 意外险
    5: 10,  # 健康险
    6: 11,  # 寿险
    7: 12,  # 医疗基金
    8: 13,  # 合计

    14: 19,  # 意外险
    15: 20,  # 健康险
    16: 21,  # 寿险
    17: 22,  # 医疗基金
    18: 23,  # 合计
}

# 核心的汇总过程
@time_cost
def merge_caled_result(upload_info: UploadInfo, result_list):
    # 1 找到对应的模板，如果没有拷贝一个新的出来
    # 2 将结果逐一填充进去
    # 返回所有可供下载的文件名称和路径的映射

    # 1. 确保年份目录存在
    target_year_dir = os.path.join(IMPORTANT_PATH, str(upload_info.year))
    os.makedirs(target_year_dir, exist_ok=True)

    # 2. 拷贝模板后填充当月数据
    fill_current_month_data(target_year_dir, upload_info, result_list)  # 填充当月数据（上传了几个核心团险数据，就计算几次当月）

    # 3. 填充内外勤人数
    if upload_info.officer_path:  # 只要上传了人员名单，那么就需要更新内外勤人数
        fill_current_month_officer(target_year_dir, upload_info.officer_path)

    # 4. 填充截止当月的全年数据
    files_map = fill_cumulative_data(target_year_dir, upload_info)  # 填充累计数据

    return files_map

# 填充当月数据
@time_cost
def fill_current_month_data(target_year_dir, upload_info: UploadInfo, result_list):
    """上传了几个团险核心数据，就需要重新计算几次当月数据
    1. 按照上传的团险核心数据的数量，拷贝模板并重命名（覆盖式）
    2. 将计算好的当月数据，填充到拷贝的模板里
    """

    # 将计算结果填充模板
    for month, result_df in zip(upload_info.upload_tuanxian_month_dict.keys(), result_list):
        this_path = os.path.join(target_year_dir, CALED_FILE.format(month=month))
        # 拷贝模板
        copy_file(TEMPLATE_PATH, this_path)

        # 使用openpyxl来操作Excel，保留原有格式
        try:
            # 加载工作簿
            wb = load_workbook(this_path)
            # 获取第一个工作表
            ws = wb.active

            # 创建分公司到行号的映射
            company_to_row = {}
            # 从第4行开始查找分公司（模板前3行是标题）
            for row in range(TEMPLATE_MIN_ROW, TEMPLATE_MAX_ROW(ws) + 1):
                company_cell = ws.cell(row=row, column=1)
                company_name = company_cell.value
                if company_name and isinstance(company_name, str):
                    company_to_row[company_name] = row

            # 将result_df中的数据填充到模板中
            for _, row_data in result_df.iterrows():
                company_name = row_data['分公司']
                if company_name in company_to_row:
                    row_num = company_to_row[company_name]
                    # 填充数据，保留原有格式
                    # 短期：<= 1年
                    for baoxian_name, col_num in TEMPLATE_BAOXIAN_COL_NUM_MAP.items():
                        ws.cell(row=row_num, column=col_num).value = row_data[baoxian_name]
                    for omit_col_num in TEMPLATE_BAOXIAN_OMIT_COL_NUM_LIST:
                        ws.cell(row=row_num, column=omit_col_num).value = 0

            # 保存工作簿
            wb.save(this_path)
            print(f"成功填充 {this_path} 的数据")
        except Exception as e:
            print(f"填充 {this_path} 的数据时出错: {str(e)}")
            continue

# 填充汇总数据
@time_cost
def fill_cumulative_data(target_year_dir, upload_info: UploadInfo):
    # 将汇总结果填充模板
    need_cal_min_month = min(upload_info.upload_tuanxian_month_dict.keys())  # 最小需要计算的月
    candidate_files = [i for i in os.listdir(target_year_dir) if i.endswith(".xlsx")]
    files = sorted(candidate_files, key=lambda x: int(x.split("月")[0]))
    files_map = {f: os.path.join(target_year_dir, f) for f in files}
    # 存储每个月的数据，用于计算后续月份的累计值
    last_month_data_dict = {}

    for file, this_path in files_map.items():  # 按顺序处理文件
        this_parts = file.split("月")
        this_month = int(this_parts[0])
        path_suffix = "月" + this_parts[1]

        # 遍历所有important中的文件（该算当月的已经保存到对应文件中了）
        # 只需要从上传的最小月开始计算汇总，但是最小的月，需要上一个月的汇总
        # if this_month < need_cal_min_month - 1:  # 需要额外多计算一个月份，用于给第一个月份做铺垫（第一个月份也需要上一个月份的汇总）
        #     continue
        # 加载当前文件
        try:
            wb = load_workbook(this_path)
            ws = wb.active
            max_row = TEMPLATE_MAX_ROW(ws)
            if this_month == 1:  # 1月的情况，汇总值就是当月值
                for row_num in range(TEMPLATE_MIN_ROW, max_row + 1):
                    for current_col_num, cumulative_col_num in TEMPLATE_CURRENT_TO_CUMULATIVE_MAP.items():
                        current_value = ws.cell(row=row_num, column=current_col_num).value or 0
                        # 确保值是数值类型
                        current_value = float(current_value) if isinstance(current_value, (int, float, str)) else 0
                        if isinstance(current_value, str):
                            try:
                                current_value = float(current_value)
                            except ValueError:
                                current_value = 0
                        ws.cell(row=row_num, column=cumulative_col_num).value = current_value

                # 保存工作簿
                wb.save(this_path)

            else:  # 非1月的情况，汇总值是上月汇总值加本月当月值
                # 找到上个月的文件
                last_month = this_month - 1
                if last_month in last_month_data_dict:
                    last_month_data = last_month_data_dict[last_month]

                    # 从第4行到第26行（包含）
                    for row_num in range(TEMPLATE_MIN_ROW, max_row + 1):
                        for cur_col_num, cum_col_num in TEMPLATE_CURRENT_TO_CUMULATIVE_MAP.items():
                            last_value = last_month_data.get(row_num, {}).get(cur_col_num) or 0
                            current_value = ws.cell(row=row_num, column=cur_col_num).value or 0

                            # 确保值是数值类型
                            last_value = float(last_value) if isinstance(last_value, (int, float, str)) else 0
                            current_value = float(current_value) if isinstance(current_value, (int, float, str)) else 0

                            # 计算并更新汇总值
                            total_value = last_value + current_value
                            ws.cell(row=row_num, column=cum_col_num).value = total_value

                    # 保存工作簿
                    wb.save(this_path)

            # 保存1月的数据供后续月份使用
            month_data = {}
            for row_num in range(TEMPLATE_MIN_ROW, max_row + 1):
                month_data[row_num] = {cur_col_num: ws.cell(row=row_num, column=cum_col_num).value or 0 for
                                       cur_col_num, cum_col_num in TEMPLATE_CURRENT_TO_CUMULATIVE_MAP.items()}

            last_month_data_dict[this_month] = month_data
        except Exception as e:
            print(f"处理 {this_path} 的汇总值时出错: {str(e)}")
            continue

    return files_map

# 填充内外勤人数
@time_cost
def fill_current_month_officer(target_year_dir, officer_path):
    if not officer_path:
        return
    """
    读取 officer_path的第一个sheet，从第5行开始，不要最后一行（是汇总）
    第一列是分公司，然后往后，每三列是一个月的数据，比如2-4列是一月，5-7月是二月
    每一个月的每一个分公司生成两个数字，内勤人数（第一列），外勤人数（第二列和第三列之和）
        比如第2月：内勤人数是第5列，外勤人数是6、7列之和

    然后遍历所有汇总月份文件，在对应的月份文件中更新内外勤人员数
    """
    # 计算各月的内外勤人数
    import pandas as pd

    try:
        # 读取人员名单Excel文件
        df_officer = pd.read_excel(officer_path, sheet_name=0, header=None)
        # 从第5行开始（索引为4），不要最后一行（汇总行）
        df_officer = df_officer.iloc[4:-1].reset_index(drop=True)
        
        # 确定分公司列名
        company_col = df_officer.columns[0]
        
        # 计算各月的内外勤人数
        month_officer_data = {}
        max_month = (len(df_officer.columns) - 1) // 3  # 每3列代表一个月的数据
        
        for month in range(1, max_month + 1):
            # 计算当前月份的数据列索引
            start_col_index = 1 + (month - 1) * 3
            
            # 提取当月数据
            month_data = {}
            for _, row in df_officer.iterrows():
                company = row[company_col]
                if pd.notna(company) and isinstance(company, str):
                    # 内勤人数（第一列）
                    office_staff = row.iloc[start_col_index] if pd.notna(row.iloc[start_col_index]) else 0
                    # 外勤人数（第二列和第三列之和）
                    field_staff1 = row.iloc[start_col_index + 1] if pd.notna(row.iloc[start_col_index + 1]) else 0
                    field_staff2 = row.iloc[start_col_index + 2] if pd.notna(row.iloc[start_col_index + 2]) else 0
                    field_staff = field_staff1 + field_staff2
                    
                    # 确保数值类型正确
                    office_staff = float(office_staff) if isinstance(office_staff, (int, float, str)) else 0
                    field_staff = float(field_staff) if isinstance(field_staff, (int, float, str)) else 0
                    
                    if isinstance(office_staff, str):
                        try:
                            office_staff = float(office_staff)
                        except ValueError:
                            office_staff = 0
                    if isinstance(field_staff, str):
                        try:
                            field_staff = float(field_staff)
                        except ValueError:
                            field_staff = 0
                    
                    month_data[company] = (office_staff, field_staff)
            
            month_officer_data[month] = month_data
    except Exception as e:
        print(f"计算各月的内外勤人数时出错: {str(e)}")
        return

    # 填充内外勤人数到每个月份的文件中
    candidate_files = [i for i in os.listdir(target_year_dir) if i.endswith(".xlsx")]
    files = sorted(candidate_files, key=lambda x: int(x.split("月")[0]))
    files_map = {f: os.path.join(target_year_dir, f) for f in files}
    
    for file_name, file_path in files_map.items():
        try:
            # 提取月份
            month_part = file_name.split("月")[0]
            if month_part.isdigit():
                month = int(month_part)
                
                # 检查该月是否有内外勤数据
                if month in month_officer_data:
                    month_data = month_officer_data[month]
                    
                    # 加载目标Excel文件
                    wb = load_workbook(file_path)
                    ws = wb.active
                    max_row = TEMPLATE_MAX_ROW(ws)
                    
                    # 创建分公司到行号的映射
                    company_to_row = {}
                    for row in range(TEMPLATE_MIN_ROW, max_row + 1):
                        company_cell = ws.cell(row=row, column=1)
                        company_name = company_cell.value
                        if company_name and isinstance(company_name, str):
                            company_to_row[company_name] = row
                    
                    # 填充内外勤人数
                    for company, (office_staff, field_staff) in month_data.items():
                        if company in company_to_row:
                            row_num = company_to_row[company]
                            # 第二列是内勤人数，第三列是外勤人数
                            ws.cell(row=row_num, column=2).value = office_staff
                            ws.cell(row=row_num, column=3).value = field_staff
                    
                    # 保存工作簿
                    wb.save(file_path)
                    print(f"成功填充 {file_path} 的内外勤人数数据")
        except Exception as e:
            print(f"填充 {file_path} 的内外勤人数数据时出错: {str(e)}")
            continue


