from lwx_project.scene.monthly_east_data.const import TEMPLATE_SHEET_NAME
import openpyxl
from openpyxl.utils import get_column_letter


def clear_and_fill_excel(df, target_file_path):
    """将df的内容逐行逐列的写入到 template_path 的 「TEMPLATE_SHEET_NAME」中
    target_file_path 的 「TEMPLATE_SHEET_NAME」中的第一行是列名，需要和df的列名对上，再填充
    :param df: 要写入的DataFrame
    :param target_file_path: 目标文件路径
    """
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(target_file_path)
        sheet = workbook[TEMPLATE_SHEET_NAME]
        
        # 获取工作表的列名（第一行）
        sheet_columns = []
        col_index = 1
        while True:
            cell_value = sheet.cell(row=1, column=col_index).value
            if cell_value is None:
                break
            sheet_columns.append(cell_value)
            col_index += 1
        
        # 确保df的列名与工作表的列名对应
        for col in df.columns:
            if col not in sheet_columns:
                raise ValueError(f"DataFrame中的列名'{col}'在工作表中不存在")
        
        # 清空现有数据行（保留第一行列名）
        # 获取当前工作表的最大行数
        max_row = sheet.max_row
        if max_row > 1:
            # 从第二行开始删除所有数据行
            sheet.delete_rows(2, max_row - 1)
        
        # 逐行逐列填充数据
        for row_idx, row in df.iterrows():
            # 从第二行开始填充数据
            excel_row = row_idx + 2
            for col in df.columns:
                # 找到该列在工作表中的位置
                col_idx = sheet_columns.index(col) + 1
                # 设置单元格值
                cell = sheet.cell(row=excel_row, column=col_idx)
                cell.value = row[col]
        
        # 保存并关闭Excel文件
        workbook.save(target_file_path)
        workbook.close()
        print(f"数据已成功填充到文件: {target_file_path}")
    except Exception as e:
        print(f"填充Excel时出错: {str(e)}")
        raise