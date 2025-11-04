import typing
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class FastExcelReader:
    """基于openpyxl 直接读取excel指定内容

    用法一：推荐
    with FastExcelReader(file_path) as fe:
        c = fe.get_excel_column_count(max_col_num=3)

    用法二
    try:
        fe = FastExcelReader(file_path)
        c = fe.get_excel_column_count(max_col_num=3)
    finally:
        fe.close()
    """

    MAX_COL_NUM = 16384  # 安全上限：Excel 最大列数为 16384 (XFD)

    def __init__(self, file_path: str, sheet_name: typing.Optional[str] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_path, read_only=True)
        self.ws = self.wb.active if self.sheet_name is None else self.wb[self.sheet_name]

    def __enter__(self):
        # 打开工作簿（只读模式，高效）
        return self  # 返回自身，供 with 块内使用

    def __exit__(self, exc_type, exc_val, exc_tb):
        # 无论是否发生异常，都关闭工作簿
        self.close()
        # 返回 False 表示不抑制异常（推荐）
        return False

    def close(self):
        self.wb.close()

    def sheets(self) -> typing.List[str]:
        return self.wb.sheetnames

    def check_sheets(self, required_sheets: typing.List[str]) -> (bool, str):
        sheets = self.sheets()
        for sheet_name in required_sheets:
            if sheet_name not in sheets:
                return False, sheet_name
        return True, ""

    def get_excel_column_count(self, max_col_num=None, row_num=1) -> int:
        """获取 Excel 文件中连续非空列的数量（从 A1 开始向右，遇到第一个空单元格即停止）。

        注意：此方法假设“列是连续的”，一旦遇到空单元格就停止计数，
        即使后续列有数据也不会被计入。

        参数:
            max_col_num (int, optional): 最多检查到第几列。
            row_num (int, optional): 用第几行作为标准计算列的个数
        返回:
            int: 第一行从左到右连续非空单元格的数量（至少为 0）。
        """
        col_num = 1
        max_col_num = max_col_num or self.MAX_COL_NUM
        try:
            while True:
                col_letter = get_column_letter(col_num)
                cell_value = self.ws[f"{col_letter}{row_num}"].value
                # None or ""
                if cell_value is None or (isinstance(cell_value, str) and len(cell_value) == 0):
                    break
                col_num += 1
                if col_num > max_col_num:
                    break
        finally:
            return col_num - 1


    def check_excel_row(self, row_num, required_value_list: list, max_col_num=None) -> (bool, list):
        max_col_num = max_col_num or self.MAX_COL_NUM
        required_value_list_copy = required_value_list[:]
        col_num = 1

        try:
            while required_value_list_copy:
                cell_value = self.get_cell_value(row_num, col_num)
                # None or ""
                if cell_value is None or (isinstance(cell_value, str) and len(cell_value) == 0):
                    pass
                else:
                    if cell_value in required_value_list_copy:
                        required_value_list_copy.remove(cell_value)
                col_num += 1
                if col_num > max_col_num:
                    break
        finally:
            return len(required_value_list_copy) == 0, required_value_list_copy

    def get_cell_value(self, row_num, col_num):
        col_letter = get_column_letter(col_num)
        cell_value = self.ws[f"{col_letter}{row_num}"].value
        return cell_value

    def check_cell_value(self, row_num, col_num, cell_value) -> bool:
        v = self.get_cell_value(row_num, col_num)
        cell_value = cell_value or ""
        if v is None:
            return isinstance(cell_value, str) and cell_value == ""

        if isinstance(v, str):
            if v.strip() == cell_value.strip():
                return True


        return False

    def posit_col_in_row_by_value(self, row_num, value, max_col_num=None) -> int:
        """根据具体内容，在特定行定位列, 返回列号col_num"""
        max_col_num = max_col_num or self.MAX_COL_NUM
        col_num = 1

        try:
            while True:
                cell_value = self.get_cell_value(row_num, col_num)
                # None or ""
                if cell_value is None or (isinstance(cell_value, str) and len(cell_value) == 0):
                    pass
                else:
                    if str(cell_value) == str(value):
                        break
                col_num += 1
                if col_num > max_col_num:
                    col_num = -1
                    break
        finally:
            return col_num

    def posit_row_in_col_by_value(self, col_num, value) -> int:
        """根据具体内容，在特定列定位行, 返回行号row_num"""
        max_row_num = self.ws.max_row
        row_num = 1

        try:
            while True:
                cell_value = self.get_cell_value(row_num, col_num)
                # None or ""
                if cell_value is None or (isinstance(cell_value, str) and len(cell_value) == 0):
                    pass
                else:
                    if str(cell_value) == str(value):
                        break
                row_num += 1
                if row_num > max_row_num:
                    col_num = -1
                    break
        finally:
            return col_num
